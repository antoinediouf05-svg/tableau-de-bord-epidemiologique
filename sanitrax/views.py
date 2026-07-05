import csv
from functools import wraps

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from . import services
from .forms import DeclarationForm
from .models import Alerte, Commune, District, Maladie, Region, Structure

# Couleurs de statut d'alerte
ST_COULEUR = {"active": "#B43A24", "prise_en_charge": "#C99A1E", "resolue": "#2E8B6F"}
ST_FOND = {"active": "#F6DAD3", "prise_en_charge": "#FAF1D6", "resolue": "#DCEFE7"}


# ════════════════════════════ Authentification & rôles ════════════════════════════
def _role(user):
    profil = getattr(user, "profil", None)
    return profil.role if profil else ("admin" if user.is_superuser else None)


def _landing(user):
    """Page d'accueil selon le rôle (les agents arrivent sur la déclaration)."""
    return "declaration" if _role(user) == "agent" else "home"


def _regions_visibles(user):
    """Périmètre géographique d'un utilisateur : un agent ne voit que SA région."""
    profil = getattr(user, "profil", None)
    if profil and profil.role == "agent" and profil.region_id:
        return Region.objects.filter(pk=profil.region_id)
    return Region.objects.all()


def role_required(*roles):
    """Restreint une vue à certains rôles (les superusers passent toujours)."""
    def decorateur(view):
        @wraps(view)
        @login_required
        def _wrapped(request, *args, **kwargs):
            if request.user.is_superuser or _role(request.user) in roles:
                return view(request, *args, **kwargs)
            raise PermissionDenied("Accès non autorisé pour votre rôle.")
        return _wrapped
    return decorateur


def connexion(request):
    """Connexion par e-mail + mot de passe (aucune inscription publique)."""
    if request.user.is_authenticated:
        return redirect(_landing(request.user))
    error = False
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        pwd = request.POST.get("password", "")
        user = None
        if email and pwd:
            compte = User.objects.filter(email__iexact=email).first()
            # On appelle toujours authenticate (même e-mail inconnu) pour que le
            # signal user_login_failed se déclenche et alimente le journal.
            user = authenticate(request, username=compte.username if compte else email,
                                password=pwd)
        if user:
            login(request, user)
            messages.success(request, f"Bienvenue, {user.get_full_name() or user.username}.")
            return redirect(request.GET.get("next") or _landing(user))
        error = True
    return render(request, "login.html", {"error": error})


def deconnexion(request):
    logout(request)
    messages.info(request, "Vous avez été déconnecté.")
    return redirect("connexion")


def _maladie_courante(request):
    """Maladie sélectionnée via ?maladie=<code> (paludisme par défaut)."""
    code = request.GET.get("maladie", "palu")
    return (Maladie.objects.filter(code=code).first()
            or Maladie.objects.filter(code="palu").first()
            or Maladie.objects.first())


def _delta(pct):
    return f"{'+' if pct >= 0 else ''}{pct} %"


def _carte_data(lignes, region_sel_code=None):
    """Liste sérialisable pour la carte Leaflet."""
    return [{
        "nom": l["region"].nom, "code": l["region"].code,
        "lat": l["region"].latitude, "lng": l["region"].longitude,
        "cases": l["cases"], "inc": l["incidence"],
        "fill": l["fill"], "label": l["label"], "niveau": l["niveau"],
        "sel": l["region"].code == region_sel_code,
    } for l in lignes]


# ──────────────────────────────── Vue d'ensemble ────────────────────────────────
@login_required
def vue_ensemble(request):
    maladie = _maladie_courante(request)
    regions = _regions_visibles(request.user)
    resume = services.resume_national(maladie, regions=regions)
    lignes = resume["lignes"]

    code_sel = request.GET.get("region")
    ligne_sel = next((l for l in lignes if l["region"].code == code_sel), lignes[0])
    part = round(ligne_sel["cases"] / resume["total_cas"] * 100) if resume["total_cas"] else 0

    # KPIs nationaux (toutes maladies, cumul) — cf. Problème 8
    tb = services.tableau_bord(regions=regions)
    kpis = [
        {"label": "Cas total", "value": tb["total"], "unit": "cas",
         "bar": "#1763C6", "delta": "cumul", "delta_couleur": "#8A99A1", "up": True},
        {"label": "Cas actifs", "value": tb["actifs"], "unit": f"· {tb['pct_actifs']} %",
         "bar": "#DD7B2E", "delta": "en cours", "delta_couleur": "#DD5A1B", "up": True},
        {"label": "Guérisons", "value": tb["gueris"], "unit": f"· {tb['pct_gueris']} %",
         "bar": "#2E8B6F", "delta": "rétablis", "delta_couleur": "#2E8B6F", "up": True},
        {"label": "Décès", "value": tb["deces"], "unit": f"· {tb['pct_deces']} %",
         "bar": "#B43A24", "delta": "létalité", "delta_couleur": "#B43A24", "up": False},
    ]

    # Activités récentes = alertes déclenchées (flux chronologique)
    activites = (Alerte.objects.filter(region__in=regions)
                 .select_related("maladie", "region").order_by("-date_detection")[:6])

    contexte = {
        "maladie": maladie,
        "maladies": Maladie.objects.all(),
        "resume": resume,
        "lignes": lignes,
        "kpis": kpis,
        "tb": tb,
        "activites": activites,
        "ligne_sel": ligne_sel,
        "part_nationale": part,
        "courbe": services.courbe(maladie),
        "alertes_recentes": (Alerte.objects.filter(region__in=regions)
                             .exclude(statut="resolue")
                             .select_related("maladie", "region")[:4]),
        "carte_data": _carte_data(lignes, ligne_sel["region"].code),
    }
    return render(request, "vue_ensemble.html", contexte)


# ──────────────────────────────── Carte & heatmap ────────────────────────────────
@login_required
def carte(request):
    maladie = _maladie_courante(request)
    lignes = services.lignes_regions(maladie, regions=_regions_visibles(request.user))
    inc_max = max((l["incidence"] for l in lignes), default=1) or 1
    for i, l in enumerate(lignes, start=1):
        l["rang"] = i
        l["pct"] = max(4, round(l["incidence"] / inc_max * 100))

    niveaux = [services.niveau_meta(n) for n in (-1, 0, 1, 2, 3)]
    contexte = {
        "maladie": maladie,
        "maladies": Maladie.objects.all(),
        "lignes": lignes,
        "carte_data": _carte_data(lignes),
        "legende": niveaux,
    }
    return render(request, "carte.html", contexte)


# ──────────────────────────────── Maladies (analyse) ────────────────────────────────
@login_required
def maladies(request):
    maladie = _maladie_courante(request)
    resume = services.resume_national(maladie, regions=_regions_visibles(request.user))
    lignes = resume["lignes"]
    inc_max = max((l["incidence"] for l in lignes), default=1) or 1
    for l in lignes:
        l["pct"] = max(4, round(l["incidence"] / inc_max * 100))

    plus_touchee = lignes[0]
    figs = [
        {"label": "Cas cumulés (16 sem.)", "value": services.cas_cumules(maladie),
         "sub": "16 sem.", "sub_couleur": "#8A99A1"},
        {"label": "Létalité", "value": f"{maladie.cfr} %".replace(".", ","),
         "sub": "cas mortels", "sub_couleur": "#8A99A1"},
        {"label": "Taux de reproduction (Rt)", "value": str(maladie.rt).replace(".", ","),
         "sub": "↑ en expansion" if maladie.rt > 1 else "↓ en recul",
         "sub_couleur": "#B43A24" if maladie.rt > 1 else "#2E8B6F"},
        {"label": "Région la plus touchée", "value": plus_touchee["region"].nom,
         "sub": f"{plus_touchee['incidence']} /100k".replace(".", ","), "sub_couleur": "#8A99A1"},
    ]
    contexte = {
        "maladie": maladie,
        "maladies": Maladie.objects.all(),
        "courbe": services.courbe(maladie),
        "demographie": services.demographie(maladie),
        "figures": figs,
        "repartition": lignes,
    }
    return render(request, "maladies.html", contexte)


# ──────────────────────────────── Centre d'alertes ────────────────────────────────
@login_required
def alertes(request):
    statut = request.GET.get("statut", "all")
    base = (Alerte.objects.filter(region__in=_regions_visibles(request.user))
            .select_related("maladie", "region"))
    comptes = {
        "all": base.count(),
        "active": base.filter(statut="active").count(),
        "prise_en_charge": base.filter(statut="prise_en_charge").count(),
        "resolue": base.filter(statut="resolue").count(),
    }
    qs = base if statut == "all" else base.filter(statut=statut)
    rows = []
    for a in qs:
        rows.append({
            "a": a,
            "meta": services.niveau_meta(a.niveau),
            "bar": min(100, round(a.ratio / 2.5 * 100)),
            "st_couleur": ST_COULEUR[a.statut],
            "st_fond": ST_FOND[a.statut],
        })
    filtres = [
        {"id": "all", "label": "Toutes", "count": comptes["all"]},
        {"id": "active", "label": "Actives", "count": comptes["active"]},
        {"id": "prise_en_charge", "label": "Prises en charge", "count": comptes["prise_en_charge"]},
        {"id": "resolue", "label": "Résolues", "count": comptes["resolue"]},
    ]
    contexte = {
        "rows": rows,
        "filtres": filtres,
        "statut_actif": statut,
    }
    return render(request, "alertes.html", contexte)


@role_required("admin", "epi", "decideur")
def prendre_en_charge(request, pk):
    """Action : passer une alerte au statut « prise en charge »."""
    alerte = get_object_or_404(Alerte, pk=pk)
    if request.method == "POST":
        alerte.statut = "prise_en_charge"
        alerte.save()
        messages.success(request, f"Alerte « {alerte.maladie.nom} — {alerte.region.nom} » prise en charge.")
    return redirect(request.POST.get("next") or "alertes")


# ──────────────────────────────── Déclaration de cas ────────────────────────────────
@role_required("admin", "agent")
def declaration(request):
    profil = getattr(request.user, "profil", None)
    agent_region = profil.region if (profil and profil.role == "agent") else None

    if request.method == "POST":
        form = DeclarationForm(request.POST, agent_region=agent_region)
        if form.is_valid():
            decl = form.save(commit=False)
            if agent_region:
                decl.region = agent_region
            decl.declare_par = request.user
            decl.save()
            services.regenerer_alertes()   # de nouvelles données peuvent déclencher des alertes
            messages.success(request, "Déclaration enregistrée et agrégée au tableau national.")
            return redirect("declaration")
    else:
        form = DeclarationForm(agent_region=agent_region, initial={
            "date_notification": services.date_reference(),
            "nombre_cas": 1,
            "statut": "suspecte",
            "maladie": Maladie.objects.filter(code="palu").first(),
        })

    districts_par_region = {}
    for d in District.objects.select_related("region"):
        districts_par_region.setdefault(str(d.region_id), []).append({"id": d.id, "nom": d.nom})

    communes_par_district = {}
    for c in Commune.objects.all():
        communes_par_district.setdefault(str(c.district_id), []).append({"id": c.id, "nom": c.nom})

    structures_par_commune = {}
    for s in Structure.objects.all():
        structures_par_commune.setdefault(str(s.commune_id), []).append(
            {"id": s.id, "nom": f"{s.nom} ({s.get_type_display()})"})

    contexte = {
        "form": form,
        "regions": Region.objects.all(),
        "districts_par_region": districts_par_region,
        "communes_par_district": communes_par_district,
        "structures_par_commune": structures_par_commune,
        "agent_region": agent_region,
    }
    return render(request, "declaration.html", contexte)


# ──────────────────────────────── Rapports ────────────────────────────────
def _periode(request):
    """Périodicité demandée : 'mois' ou 'semaine' (défaut)."""
    p = request.GET.get("periode", "semaine")
    return "mois" if p == "mois" else "semaine"


@login_required
def rapports(request):
    periode = _periode(request)
    debut, terme, label = services.plage_periode(periode)
    contexte = {
        "periode": periode,
        "periode_label": label,
        "debut": debut,
        "terme": terme,
    }
    return render(request, "rapports.html", contexte)


# En-têtes du bulletin (toutes maladies × toutes régions)
BULLETIN_COLS = ["Région", "Maladie", "Cas", "Incidence /100k", "Seuil", "Niveau"]


def _bulletin_lignes(periode):
    lignes, contexte = services.bulletin_donnees(periode)
    rows = [[l["region"], l["maladie"], l["cas"], l["incidence"], l["seuil"], l["niveau"]]
            for l in lignes]
    return rows, contexte


@login_required
def export_csv(request):
    """Export CSV brut de l'agrégation (toutes maladies × régions) sur la période."""
    periode = _periode(request)
    rows, ctx = _bulletin_lignes(periode)
    reponse = HttpResponse(content_type="text/csv; charset=utf-8")
    reponse["Content-Disposition"] = f'attachment; filename="SNSE_bulletin_{periode}.csv"'
    reponse.write("﻿")  # BOM pour Excel
    writer = csv.writer(reponse, delimiter=";")
    writer.writerow(BULLETIN_COLS)
    writer.writerows(rows)
    return reponse


@login_required
def export_excel(request):
    """Bulletin épidémiologique au format Excel (.xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    periode = _periode(request)
    rows, ctx = _bulletin_lignes(periode)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bulletin"
    ws.append([f"Bulletin épidémiologique {ctx['periode']} — "
               f"{ctx['debut']:%d/%m/%Y} au {ctx['terme']:%d/%m/%Y}"])
    ws.append([])
    ws.append(BULLETIN_COLS)
    entete = ws[3]
    for cell in entete:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1763C6")
    for r in rows:
        ws.append(r)
    for i, col in enumerate(BULLETIN_COLS, start=1):
        ws.column_dimensions[chr(64 + i)].width = max(12, len(col) + 4)

    reponse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    reponse["Content-Disposition"] = f'attachment; filename="SNSE_bulletin_{periode}.xlsx"'
    wb.save(reponse)
    return reponse


@login_required
def export_pdf(request):
    """Bulletin épidémiologique au format PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    periode = _periode(request)
    rows, ctx = _bulletin_lignes(periode)

    reponse = HttpResponse(content_type="application/pdf")
    reponse["Content-Disposition"] = f'attachment; filename="SNSE_bulletin_{periode}.pdf"'

    doc = SimpleDocTemplate(reponse, pagesize=landscape(A4),
                            topMargin=28, bottomMargin=28, leftMargin=28, rightMargin=28)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Bulletin épidémiologique national — Sanitrax / SNSE", styles["Title"]),
        Paragraph(f"Période : {ctx['periode']} · {ctx['debut']:%d/%m/%Y} au "
                  f"{ctx['terme']:%d/%m/%Y}", styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [BULLETIN_COLS] + [[str(c) for c in r] for r in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1763C6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
    ]))
    elements.append(table)
    doc.build(elements)
    return reponse
